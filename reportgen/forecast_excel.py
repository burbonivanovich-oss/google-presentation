"""Сборка Excel-прогноза 2026 по Маркету и ОФД.

Источники:
  * Царь свод все разрезы.xlsx — план/факт помесячно по проектам
        ОФД = п453, Маркет = п470 + п47101
  * data (15).xlsx — рекламный факт (целевые лиды) помесячно по продукту
  * дашборды Контекст (ad_dashboards.py) — расход / оплаты / CPL

На выходе .xlsx с листами:
  Сводка · Маркет (выручка) · Маркет (оплаты) · ОФД (выручка) · ОФД (оплаты) ·
  Сравнение методов · Реклама — эффективность · Методология · Данные (вход)
загружается в указанную папку Drive.
"""
from __future__ import annotations

import io
import re

import pandas as pd
from googleapiclient.http import MediaIoBaseUpload

from .aggregator import (
    PF_QTY_FACT, PF_QTY_PLAN, PF_REV_FACT, PF_REV_PLAN,
    MONTHS_RU, _clean_plan_fact, _download_xlsx, _find, _list_xlsx, _to_num,
)
from .ad_dashboards import AD, MONTHS_ORDER
from .drive import DriveClient
from .forecast_models import MONTHS, forecast_series

PRODUCTS = {"ОФД": ["п453"], "Маркет": ["п470", "п47101"]}
PRODUCT_UCHETA = {"ОФД": "Контур.ОФД", "Маркет": "Контур.Маркет"}
MONTHS_SHORT = {1: "янв", 2: "фев", 3: "мар", 4: "апр", 5: "май", 6: "июн",
                7: "июл", 8: "авг", 9: "сен", 10: "окт", 11: "ноя", 12: "дек"}


def _proj_code(s) -> str:
    m = re.match(r"\s*(п?\d+)", str(s).strip().lower())
    return m.group(1) if m else str(s).strip().lower()


# ── чтение источников ─────────────────────────────────────────────

def _plan_fact_by_project(df: pd.DataFrame, projects: list[str], col: str,
                          year: int) -> dict[int, float]:
    codes = {_proj_code(p) for p in projects}
    mask = df["Проект"].apply(lambda s: _proj_code(s) in codes)
    sub = df[mask & (df["_year"] == year)]
    out = {}
    for m in MONTHS:
        out[m] = float(sub[sub["_month"] == m][col].sum()) if col in sub else 0.0
    return out


def _ad_attributed(d15: pd.DataFrame, product_uchet: str) -> dict:
    """data(15): период в 1-й колонке ('янв 2025'), столбцы Продукт учета,
    Оплаты, Выручка. Вернём помесячно выручку и оплаты по продукту."""
    df = d15.copy()
    # найдём строку-заголовок
    hdr = None
    for i in range(min(len(df), 6)):
        if any(str(v).strip() == "Выручка" for v in df.iloc[i].values):
            hdr = i
            break
    if hdr is not None:
        df.columns = [str(v).strip() for v in df.iloc[hdr].values]
        df = df.iloc[hdr + 1:].reset_index(drop=True)
    period_col = df.columns[0]
    prod_col = next((c for c in df.columns if "Продукт уч" in str(c)), None)
    rev = {y_m: 0.0 for y_m in MONTHS_ORDER}
    opl = {y_m: 0.0 for y_m in MONTHS_ORDER}
    if prod_col is None or "Выручка" not in df.columns:
        return {"rev": rev, "opl": opl}
    for _, r in df.iterrows():
        if str(r.get(prod_col)).strip() != product_uchet:
            continue
        mm = re.match(r"([а-я]{3})\s+(\d{4})", str(r[period_col]).strip().lower())
        if not mm:
            continue
        month = {v: k for k, v in MONTHS_SHORT.items()}.get(mm.group(1))
        yr = int(mm.group(2))
        if month is None or (yr, month) not in rev:
            continue
        rev[(yr, month)] += _to_num(r.get("Выручка"))
        opl[(yr, month)] += _to_num(r.get("Оплаты"))
    return {"rev": rev, "opl": opl}


def build_forecast_data(drive: DriveClient, folder_id: str, completed: int) -> dict:
    files = _list_xlsx(drive, folder_id)
    svod = _find(files, [r"Царь свод", r"свод все разрезы"])
    d15f = _find(files, [r"data \(15\)"])
    if not svod:
        raise RuntimeError("Не найден Царь свод все разрезы.xlsx")
    df = _clean_plan_fact(_download_xlsx(drive, svod["id"]))
    d15 = _download_xlsx(drive, d15f["id"]) if d15f else None

    result = {"completed": completed, "products": {}}
    for name, projects in PRODUCTS.items():
        rev_25 = _plan_fact_by_project(df, projects, PF_REV_FACT, 2025)
        rev_26 = _plan_fact_by_project(df, projects, PF_REV_FACT, 2026)
        plan_rev = _plan_fact_by_project(df, projects, PF_REV_PLAN, 2026)
        qty_25 = _plan_fact_by_project(df, projects, PF_QTY_FACT, 2025)
        qty_26 = _plan_fact_by_project(df, projects, PF_QTY_FACT, 2026)
        plan_qty = _plan_fact_by_project(df, projects, PF_QTY_PLAN, 2026)

        result["products"][name] = {
            "projects": projects,
            "rev_2025": rev_25, "rev_2026": rev_26, "plan_rev": plan_rev,
            "qty_2025": qty_25, "qty_2026": qty_26, "plan_qty": plan_qty,
            "fc_rev": forecast_series(rev_25, rev_26, plan_rev, completed),
            "fc_qty": forecast_series(qty_25, qty_26, plan_qty, completed),
            "ad_attr": _ad_attributed(d15, PRODUCT_UCHETA[name]) if d15 is not None
                       else {"rev": {}, "opl": {}},
        }
    return result


# ── запись Excel ──────────────────────────────────────────────────

def write_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ORANGE = "F28C0D"
    DARK = "1D2430"
    GRAY = "5E6670"
    LIGHT = "FFF4E2"
    GREEN = "2F9E44"
    RED = "E03131"

    hdr_fill = PatternFill("solid", fgColor=ORANGE)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color=DARK)
    sub_font = Font(italic=True, size=10, color=GRAY)
    bold = Font(bold=True)
    money = "#,##0 ₽"
    pct = "0.0%"
    thin = Side(style="thin", color="E7E9EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    completed = data["completed"]
    comp_name = MONTHS_RU[completed]

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border

    def autosize(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── 1. Сводка ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Прогноз выполнения плана 2026 — Маркет и ОФД"
    ws["A1"].font = title_font
    ws["A2"] = (f"Факт по {comp_name} включительно · модели: plan-pacing, "
                f"сезонный YoY, ETS (statsmodels), регрессия (sklearn) · "
                f"ансамбль и сценарии низ/база/верх")
    ws["A2"].font = sub_font
    autosize(ws, [34, 16, 16, 16, 16, 16, 16])

    row = 4
    for metric, key in (("ВЫРУЧКА, ₽", "fc_rev"), ("ОПЛАТЫ, шт", "fc_qty")):
        ws.cell(row=row, column=1, value=metric).font = Font(bold=True, size=12, color=ORANGE)
        row += 1
        headers = ["Продукт / показатель", "Маркет", "ОФД"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, 3)
        row += 1
        m = data["products"]["Маркет"][key]
        o = data["products"]["ОФД"][key]
        is_money = key == "fc_rev"
        fmt = money if is_money else "#,##0"

        def put(label, vm, vo, f=fmt, bold_row=False, color=None):
            nonlocal row
            ws.cell(row=row, column=1, value=label)
            cm = ws.cell(row=row, column=2, value=vm)
            co = ws.cell(row=row, column=3, value=vo)
            for cc in (cm, co):
                cc.number_format = f
                if bold_row:
                    cc.font = bold
                if color:
                    cc.font = Font(bold=bold_row, color=color)
            if bold_row:
                ws.cell(row=row, column=1).font = bold
            row += 1

        put("Годовой план", m.annual_plan, o.annual_plan)
        put(f"План YTD (янв–{MONTHS_SHORT[completed]})", m.ytd_plan, o.ytd_plan)
        put(f"Факт YTD (янв–{MONTHS_SHORT[completed]})", m.ytd_fact, o.ytd_fact, bold_row=True)
        put("Выполнение YTD, %", m.fulfil_ytd / 100, o.fulfil_ytd / 100, f=pct)
        put("YoY к 2025, %", m.yoy / 100, o.yoy / 100, f=pct,
            color=GREEN if (m.yoy >= 0 and o.yoy >= 0) else None)
        put("Прогноз года — НИЗ", m.year_low, o.year_low)
        put("Прогноз года — БАЗА", m.year_base, o.year_base, bold_row=True)
        put("Прогноз года — ВЕРХ", m.year_high, o.year_high)
        put("Выполнение плана — НИЗ, %", m.pct_plan_low / 100, o.pct_plan_low / 100, f=pct)
        put("Выполнение плана — БАЗА, %", m.pct_plan_base / 100, o.pct_plan_base / 100,
            f=pct, bold_row=True)
        put("Выполнение плана — ВЕРХ, %", m.pct_plan_high / 100, o.pct_plan_high / 100, f=pct)
        # вердикт
        def verdict(p):
            if p >= 100: return "план будет выполнен"
            if p >= 95: return "на грани плана"
            if p >= 85: return "недовыполнение"
            return "существенный недобор"
        ws.cell(row=row, column=1, value="Вывод (база)")
        ws.cell(row=row, column=2, value=verdict(m.pct_plan_base))
        ws.cell(row=row, column=3, value=verdict(o.pct_plan_base))
        ws.cell(row=row, column=1).font = bold
        row += 2

    for r in range(4, row):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and cell.border.left.style is None:
                cell.border = border

    # ── помесячные листы выручки/оплат ───────────────────────────
    def month_sheet(title, prod, key, is_money):
        ws = wb.create_sheet(title)
        p = data["products"][prod]
        fc = p[key]
        plan = p["plan_rev"] if is_money else p["plan_qty"]
        f25 = p["rev_2025"] if is_money else p["qty_2025"]
        f26 = p["rev_2026"] if is_money else p["qty_2026"]
        fmt = money if is_money else "#,##0"
        ws["A1"] = f"{prod} — {'выручка' if is_money else 'оплаты'} 2026, помесячно"
        ws["A1"].font = title_font
        ws["A2"] = (f"Проекты {', '.join(p['projects'])} · факт по {comp_name}, "
                    f"далее прогноз (база) и коридор низ/верх")
        ws["A2"].font = sub_font
        head = ["Месяц", "План 2026", "Факт 2025", "Факт 2026",
                "Прогноз база", "Прогноз низ", "Прогноз верх"]
        hr = 4
        for c, h in enumerate(head, 1):
            ws.cell(row=hr, column=c, value=h)
        style_header(ws, hr, len(head))
        for i, mth in enumerate(MONTHS):
            r = hr + 1 + i
            ws.cell(row=r, column=1, value=MONTHS_RU[mth])
            ws.cell(row=r, column=2, value=plan.get(mth, 0))
            ws.cell(row=r, column=3, value=f25.get(mth, 0))
            if mth <= completed:
                ws.cell(row=r, column=4, value=f26.get(mth, 0))
            else:
                ws.cell(row=r, column=5, value=fc.base.get(mth, 0))
                ws.cell(row=r, column=6, value=fc.low.get(mth, 0))
                ws.cell(row=r, column=7, value=fc.high.get(mth, 0))
            for c in range(2, 8):
                cc = ws.cell(row=r, column=c)
                cc.number_format = fmt
                cc.border = border
            ws.cell(row=r, column=1).border = border
        # итог
        tr = hr + 1 + len(MONTHS)
        ws.cell(row=tr, column=1, value="ИТОГО год").font = bold
        ws.cell(row=tr, column=2, value=fc.annual_plan).number_format = fmt
        ws.cell(row=tr, column=4, value=fc.ytd_fact).number_format = fmt
        ws.cell(row=tr, column=5, value=fc.year_base).number_format = fmt
        ws.cell(row=tr, column=6, value=fc.year_low).number_format = fmt
        ws.cell(row=tr, column=7, value=fc.year_high).number_format = fmt
        for c in range(1, 8):
            ws.cell(row=tr, column=c).font = bold
            ws.cell(row=tr, column=c).fill = PatternFill("solid", fgColor=LIGHT)
        autosize(ws, [12, 15, 15, 15, 15, 15, 15])

        chart = LineChart()
        chart.title = f"{prod}: план / факт / прогноз"
        chart.height, chart.width = 9, 20
        data_ref = Reference(ws, min_col=2, max_col=7, min_row=hr, max_row=hr + 12)
        cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=hr + 12)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "I4")
        return ws

    month_sheet("Маркет (выручка)", "Маркет", "fc_rev", True)
    month_sheet("Маркет (оплаты)", "Маркет", "fc_qty", False)
    month_sheet("ОФД (выручка)", "ОФД", "fc_rev", True)
    month_sheet("ОФД (оплаты)", "ОФД", "fc_qty", False)

    # ── сравнение методов ────────────────────────────────────────
    ws = wb.create_sheet("Сравнение методов")
    ws["A1"] = "Прогноз года по методам (выручка), ₽"
    ws["A1"].font = title_font
    ws["A2"] = "Ансамбль (база) — взвешенное среднее: pacing 0.40 · YoY 0.25 · ETS 0.15 · регрессия 0.20"
    ws["A2"].font = sub_font
    method_names = {"plan_pacing": "Plan-pacing", "seasonal_yoy": "Сезонный YoY",
                    "ets_holt": "ETS / Holt", "linreg_season": "Регрессия+сезон"}
    hr = 4
    ws.cell(row=hr, column=1, value="Метод")
    ws.cell(row=hr, column=2, value="Маркет")
    ws.cell(row=hr, column=3, value="ОФД")
    style_header(ws, hr, 3)
    r = hr + 1
    for mk in ["plan_pacing", "seasonal_yoy", "ets_holt", "linreg_season"]:
        ws.cell(row=r, column=1, value=method_names[mk])
        for c, prod in ((2, "Маркет"), (3, "ОФД")):
            v = data["products"][prod]["fc_rev"].year_by_method.get(mk)
            cell = ws.cell(row=r, column=c, value=v if v is not None else "—")
            cell.number_format = money
            cell.border = border
        ws.cell(row=r, column=1).border = border
        r += 1
    ws.cell(row=r, column=1, value="АНСАМБЛЬ (база)").font = bold
    for c, prod in ((2, "Маркет"), (3, "ОФД")):
        cell = ws.cell(row=r, column=c, value=data["products"][prod]["fc_rev"].year_base)
        cell.number_format = money
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor=LIGHT)
    autosize(ws, [22, 18, 18])

    # ── реклама — эффективность ──────────────────────────────────
    ws = wb.create_sheet("Реклама — эффективность")
    ws["A1"] = "Контекстная реклама: расход / оплаты / CPO / CPL"
    ws["A1"].font = title_font
    ws["A2"] = ("Источник: дашборды Контекст (распознано с экрана, суммы сверены с «Итого»). "
                "Выручка по целевым лидам — из data (15). Июнь 2026 — неполный.")
    ws["A2"].font = sub_font
    head = ["Месяц", "Маркет расход", "Маркет опл.", "Маркет CPO", "Маркет CPL", "Маркет выр.лиды",
            "ОФД расход", "ОФД опл.", "ОФД CPO", "ОФД CPL", "ОФД выр.лиды"]
    hr = 4
    for c, h in enumerate(head, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(head))
    for i, (y, mth) in enumerate(MONTHS_ORDER):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=f"{MONTHS_SHORT[mth]} {y}")
        for base_c, prod in ((2, "Маркет"), (7, "ОФД")):
            spend = AD[prod]["spend"].get((y, mth), 0)
            opl = AD[prod]["oplaty"].get((y, mth), 0)
            cpl = AD[prod]["cpl"].get((y, mth), 0)
            cpo = spend / opl if opl else 0
            adrev = data["products"][prod]["ad_attr"]["rev"].get((y, mth), 0)
            ws.cell(row=r, column=base_c, value=spend).number_format = money
            ws.cell(row=r, column=base_c + 1, value=opl)
            ws.cell(row=r, column=base_c + 2, value=round(cpo)).number_format = money
            ws.cell(row=r, column=base_c + 3, value=cpl).number_format = money
            ws.cell(row=r, column=base_c + 4, value=round(adrev)).number_format = money
        for c in range(1, len(head) + 1):
            ws.cell(row=r, column=c).border = border
    autosize(ws, [10] + [14] * 10)

    # ── методология ──────────────────────────────────────────────
    ws = wb.create_sheet("Методология")
    ws.column_dimensions["A"].width = 120
    lines = [
        ("Методология прогноза 2026 — Маркет и ОФД", title_font),
        ("", None),
        ("1. ИСТОЧНИКИ ДАННЫХ", bold),
        ("• Царь свод все разрезы.xlsx — помесячный план и факт по проектам.", None),
        ("    Маркет = проекты п470 + п47101; ОФД = проект п453 (п45302 — техника/ФН/услуги внедрения — исключён).", None),
        ("• data (15).xlsx — рекламный факт (целевые лиды по продукту): выручка и оплаты по месяцам.", None),
        ("• Дашборды Контекст (PNG) — расход, оплаты, CPL по месяцам (распознано с экрана, суммы сверены с «Итого»).", None),
        ("", None),
        ("2. ГОРИЗОНТ И ОБУЧАЮЩАЯ ВЫБОРКА", bold),
        (f"• Факт известен по {comp_name} 2026 включительно (YTD). Прогнозируются оставшиеся месяцы до декабря 2026.", None),
        ("• Обучающий ряд: янв 2025 – последний завершённый месяц 2026 (~17 точек). Июнь 2026 неполный — в обучение не берётся.", None),
        ("", None),
        ("3. МОДЕЛИ (считаются независимо, затем ансамбль)", bold),
        ("• Plan-pacing: оставшийся месяц = план месяца × (факт YTD ÷ план YTD).", None),
        ("    Самый бизнес-обоснованный — план уже содержит сезонность, а текущий темп выполнения переносится на остаток.", None),
        ("• Сезонный YoY: оставшийся месяц = факт того же месяца 2025 × (1 + g), g — прирост YTD-2026 к тому же периоду 2025.", None),
        ("• ETS / Holt (statsmodels ExponentialSmoothing): аддитивный демпфированный тренд; сезонность period=12, если ≥24 точек.", None),
        ("• Регрессия + сезонные дамми (scikit-learn LinearRegression): линейный тренд по времени + 11 месячных индикаторов.", None),
        ("", None),
        ("4. АНСАМБЛЬ И СЦЕНАРИИ", bold),
        ("• База = взвешенное среднее доступных методов: plan-pacing 0.40, сезонный YoY 0.25, ETS 0.15, регрессия 0.20.", None),
        ("• Коридор: НИЗ = минимум по методам в каждом месяце, ВЕРХ = максимум. Сумма за год даёт сценарии низ/база/верх.", None),
        ("• Отрицательные прогнозы обнуляются.", None),
        ("", None),
        ("5. ВЫПОЛНЕНИЕ ПЛАНА", bold),
        ("• Прогноз года = факт YTD + Σ прогноза остатка. Выполнение плана = прогноз года ÷ годовой план × 100%.", None),
        ("• Вывод: ≥100% — план будет выполнен; 95–100% — на грани; 85–95% — недовыполнение; <85% — существенный недобор.", None),
        ("", None),
        ("6. ОГРАНИЧЕНИЯ", bold),
        ("• Короткий ряд (~1.5 года) — сезонные модели работают на пределе; ставка сделана на plan-pacing и YoY.", None),
        ("• Рекламные расход/CPL сняты с дашбордов распознаванием — возможны мелкие неточности (итоги сверены).", None),
        ("• Прогноз не учитывает разовые акции, изменения цен и план-факт корректировки после даты выгрузки.", None),
    ]
    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── данные (вход) ────────────────────────────────────────────
    ws = wb.create_sheet("Данные (вход)")
    ws["A1"] = "Входные помесячные ряды (план/факт по проектам), ₽"
    ws["A1"].font = title_font
    hr = 3
    cols = ["Продукт", "Показатель", "Год"] + [MONTHS_SHORT[m] for m in MONTHS]
    for c, h in enumerate(cols, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(cols))
    r = hr + 1
    for prod in ("Маркет", "ОФД"):
        p = data["products"][prod]
        for label, series in (("План 2026", p["plan_rev"]), ("Факт 2025", p["rev_2025"]),
                              ("Факт 2026", p["rev_2026"])):
            ws.cell(row=r, column=1, value=prod)
            ws.cell(row=r, column=2, value=label)
            ws.cell(row=r, column=3, value=2026 if "2026" in label else 2025)
            for c, m in enumerate(MONTHS, 4):
                ws.cell(row=r, column=c, value=round(series.get(m, 0))).number_format = "#,##0"
            r += 1
    autosize(ws, [10, 12, 7] + [11] * 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_xlsx(drive: DriveClient, folder_id: str, filename: str, content: bytes) -> str:
    media = MediaIoBaseUpload(
        io.BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    meta = {"name": filename, "parents": [folder_id]}
    f = drive._drive.files().create(  # noqa: SLF001
        body=meta, media_body=media, fields="id,webViewLink",
        supportsAllDrives=True,
    ).execute()
    return f.get("webViewLink") or f.get("id")
